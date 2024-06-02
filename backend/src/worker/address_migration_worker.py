import csv
import logging
import os
import subprocess
import zipfile
from datetime import timedelta, datetime
from threading import Thread
from time import sleep

import requests
from openpyxl import load_workbook

from src.database.generic_dal import get_last_migration, add_migration, delete_all, insert_bulk
from src.database.models.addresses import Addresses
from src.database.provider import db_provider


class AddressMigrationWorker:
    def __init__(self):
        self.logger = logging.getLogger("AddressMigrationWorker")
        self.db_provider = db_provider
        self.migration_thread = Thread(target=self._migrate_loop)
        self.sleep_time = 10
        self.running = True
        self.service_url = 'https://geoportal.wroclaw.pl/www/emuia/Adresy.zip'
        self.migration_interval = timedelta(days=7)
        self.last_migration = get_last_migration()
        self.zip_path = 'Adresy.zip'
        self.destination = 'Adresy'
        self.csv_path = 'addresses.csv'
        self.table_name = 'Addresses'

    def start(self):
        self.logger.info("Starting address migration worker")
        self.migration_thread.start()

    def stop(self):
        self.logger.info("Stopping address migration worker")
        self.running = False
        self.migration_thread.join()

    def _migrate_loop(self):
        while self.running:
            if self.last_migration is None or datetime.now() - self.last_migration.date > self.migration_interval:
                self.logger.info("Starting address migration...")
                self.__migrate()
                self.logger.info("Address migration finished")
            sleep(self.sleep_time)

    def __migrate(self):
        request = requests.get(self.service_url)
        if request.status_code != 200:
            self.logger.error("Failed to download addresses file")
            return

        self.logger.info("Processing new addresses file...")
        with open(self.zip_path, 'wb') as f:
            f.write(request.content)

        with zipfile.ZipFile(self.zip_path, 'r') as zip_ref:
            zip_ref.extractall(self.destination)

        filenames = os.listdir(self.destination)
        if not filenames or len(filenames) == 0:
            self.logger.error("No files in the extracted zip")
            return
        xlsx_path = os.path.join(self.destination, filenames[0])
        try:
            self.logger.info("Deleting old addresses...")
            delete_all(Addresses)
            self.__parse_and_copy(xlsx_path)
            add_migration()
        except Exception as e:
            self.logger.error(f"Failed to migrate addresses")

    def __parse_and_copy(self, file_path):
        self.logger.info(f"Parsing xlsx data file...")
        try:
            labels = {2: 'street', 3: 'number', 4: 'postal_code', 6: 'gus_terc', 7: 'settlement'}
            data = []
            workbook = load_workbook(file_path)
            sheet = workbook.active
            header = False
            for row in sheet.iter_rows(values_only=True):
                if not header:
                    header = True
                    continue
                modified_row = {labels[i]: row[i] for i in range(len(row)) if i not in [0, 1, 5]}
                data.append(modified_row)
        except Exception as e:
            self.logger.error(f"Failed to parse XLSX file: {e}")
            raise

        self.logger.info(f"Importing {len(data)} rows to db...")
        try:
            insert_bulk(Addresses, data)
        except Exception as e:
            self.logger.error(f"Failed to import db data: {e}")
            raise


migration_worker = AddressMigrationWorker()
